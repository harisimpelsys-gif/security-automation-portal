#!/usr/bin/env python3
"""
WFH/WFO Roster Generator

Flow:
1. On run, asks for:
   - Year
   - Month (number or name, e.g. "12" or "December")
2. Asks if you want to add new members (yes/no).
   - Employee list is stored in employees.txt (one name per line).
3. Generates an Excel roster for that month:
   - Weekdays only (Mon–Fri)
   - Title row (e.g. "December 2025")
   - Week labels: 1st Week, 2nd Week, 3rd Week, etc.
   - Day names row
   - Date row (dd-MMM)
   - Employee names in column A
   - WFO/WFH cells left empty for manual filling.
"""

import os
import calendar
from datetime import datetime
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


EMP_FILE = "employees.txt"  # one employee name per line


def load_employees(path: str = EMP_FILE):
    """Load employees from a text file. If not existing, ask user to enter base list."""
    if not os.path.exists(path):
        print(f"Employee file '{path}' not found.")
        print("Please enter initial employee names. Leave blank when done.")
        employees = []
        while True:
            name = input("Enter employee name (blank to finish): ").strip()
            if not name:
                break
            employees.append(name)
        if not employees:
            raise ValueError("No employees provided.")
        save_employees(employees, path)
        return employees

    with open(path, "r", encoding="utf-8") as f:
        employees = [line.strip() for line in f.readlines() if line.strip()]

    if not employees:
        raise ValueError(f"Employee file '{path}' is empty.")
    return employees


def save_employees(employees, path: str = EMP_FILE):
    """Save employees to text file (one per line)."""
    with open(path, "w", encoding="utf-8") as f:
        for name in employees:
            f.write(name + "\n")


def ask_new_members(employees):
    """Ask user if they want to add new members; update list if yes."""
    ans = input("Do you want to add new member(s)? (y/n): ").strip().lower()
    if ans not in ("y", "yes"):
        return employees

    print("Enter new members. Leave blank when done.")
    while True:
        name = input("New member name (blank to finish): ").strip()
        if not name:
            break
        if name not in employees:
            employees.append(name)
        else:
            print(f"'{name}' already in the list, skipping.")

    return employees


def parse_month_input(year: int):
    """Ask user for month and return month number (1–12)."""
    while True:
        month_in = input("Enter month (number like 12 or name like December): ").strip()
        if not month_in:
            print("Month cannot be empty.")
            continue

        # Try numeric
        if month_in.isdigit():
            m = int(month_in)
            if 1 <= m <= 12:
                return m
            print("Invalid month number. Please enter between 1 and 12.")
            continue

        # Try name
        month_in_lower = month_in.lower()
        month_map = {calendar.month_name[i].lower(): i for i in range(1, 13)}
        month_map.update({calendar.month_abbr[i].lower(): i for i in range(1, 13)})

        if month_in_lower in month_map:
            return month_map[month_in_lower]

        print("Invalid month input. Try again (e.g. '12', 'Dec', or 'December').")


def ordinal(n: int) -> str:
    """Return ordinal string: 1 -> '1st', 2 -> '2nd' etc."""
    if 10 <= n % 100 <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"


def generate_roster(year: int, month: int, employees):
    """Generate the WFH/WFO roster Excel file for the given month and employees."""
    # All weekdays in the month (Mon–Fri)
    num_days = calendar.monthrange(year, month)[1]
    dates = [
        datetime(year, month, day)
        for day in range(1, num_days + 1)
        if datetime(year, month, day).weekday() < 5  # 0=Mon .. 4=Fri
    ]

    if not dates:
        raise ValueError("No weekdays in this month? Something is wrong.")

    wb = Workbook()
    ws = wb.active

    month_name = calendar.month_name[month]
    sheet_name = f"{calendar.month_abbr[month]}-{str(year)[-2:]}"
    ws.title = sheet_name

    total_cols = len(dates) + 1  # +1 for Resource column

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"{month_name} {year}"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    # Week groups (based on ISO week number order of appearance)
    week_groups = OrderedDict()
    for col_index, d in enumerate(dates, start=2):  # columns start at 2 (B)
        iso_week = d.isocalendar()[1]
        if iso_week not in week_groups:
            week_groups[iso_week] = [col_index]
        else:
            week_groups[iso_week].append(col_index)

    # Row indices
    row_week = 2       # Week labels
    row_dayname = 3    # Day names
    row_date = 4       # Dates
    row_emp_start = 5  # Employee names

    # Week labels row
    for i, (iso_week, cols) in enumerate(week_groups.items(), start=1):
        start_col = cols[0]
        end_col = cols[-1]
        start_letter = get_column_letter(start_col)
        end_letter = get_column_letter(end_col)
        ws.merge_cells(f"{start_letter}{row_week}:{end_letter}{row_week}")

        label = f"{ordinal(i)} Week"
        cell = ws.cell(row=row_week, column=start_col)
        cell.value = label
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Day name + date rows
    ws.cell(row=row_dayname, column=1).value = "Resource"
    ws.cell(row=row_dayname, column=1).font = Font(bold=True)
    ws.cell(row=row_dayname, column=1).alignment = Alignment(horizontal="center", vertical="center")

    for col_index, d in enumerate(dates, start=2):
        # Day name (e.g. Monday)
        c_day = ws.cell(row=row_dayname, column=col_index)
        c_day.value = d.strftime("%A")
        c_day.alignment = Alignment(horizontal="center", vertical="center")

        # Date (e.g. 01-Dec)
        c_date = ws.cell(row=row_date, column=col_index)
        c_date.value = d.strftime("%d-%b")
        c_date.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions[get_column_letter(col_index)].width = 12

    # Employee names
    for i, emp in enumerate(employees):
        r = row_emp_start + i
        c_emp = ws.cell(row=r, column=1)
        c_emp.value = emp
        c_emp.alignment = Alignment(horizontal="left", vertical="center")

    # Header styling
    header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    for col in range(1, total_cols + 1):
        for r in (row_week, row_dayname, row_date):
            ws.cell(row=r, column=col).fill = header_fill

    # Save file
    out_filename = f"SecOps_WFO_Tracker_{sheet_name}.xlsx"
    wb.save(out_filename)
    return out_filename


def main():
    print("=== WFH/WFO Roster Generator ===")

    # Year input
    while True:
        year_str = input("Enter year (e.g. 2025): ").strip()
        if not year_str.isdigit():
            print("Invalid year. Try again.")
            continue
        year = int(year_str)
        if year < 2000 or year > 2100:
            print("Year out of range. Try again.")
            continue
        break

    month = parse_month_input(year)

    # Load + optionally update employees
    employees = load_employees()
    employees = ask_new_members(employees)
    save_employees(employees)

    # Generate file
    out_file = generate_roster(year, month, employees)
    print(f"Roster generated: {out_file}")


if __name__ == "__main__":
    main()
