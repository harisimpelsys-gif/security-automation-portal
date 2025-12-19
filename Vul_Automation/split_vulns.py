#!/usr/bin/env python3
"""
split_vulns.py
"""

import argparse
import pathlib
import re
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

AGE_COLOR_MAP = {
    "0-30": "C6EFCE",
    "31-60": "FFEB9C",
    "61-90": "F4B084",
    "90+": "FFC7CE",
}

AGE_KEYWORDS = {
    "0": "C6EFCE",
    "30": "C6EFCE",
    "31": "FFEB9C",
    "60": "FFEB9C",
    "61": "F4B084",
    "90": "FFC7CE",
}

INVALID_SHEET_CHARS = r'[:\\/?*\[\]]'

def sanitize_sheet_name(name: str) -> str:
    name = re.sub(INVALID_SHEET_CHARS, "_", str(name)).strip()
    return name[:31] if name else "Sheet"

def pick_age_color(cell_val):
    if pd.isna(cell_val):
        return None
    s = str(cell_val).strip()
    if s in AGE_COLOR_MAP:
        return AGE_COLOR_MAP[s]
    for k, v in AGE_COLOR_MAP.items():
        if k.replace("-", "") in s.replace(" ", "").replace("-", ""):
            return v
    for kw, color in AGE_KEYWORDS.items():
        if kw in s:
            return color
    return None

def style_workbook(path: pathlib.Path, age_col_name="Age In Days"):
    wb = load_workbook(path)
    for ws in wb.worksheets:
        for c in range(1, ws.max_column + 1):
            ws.cell(row=1, column=c).font = Font(bold=True)

        age_col_idx = None
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(row=1, column=c).value).lower() == age_col_name.lower():
                age_col_idx = c
                break

        if age_col_idx:
            for r in range(2, ws.max_row + 1):
                color = pick_age_color(ws.cell(row=r, column=age_col_idx).value)
                if color:
                    ws.cell(row=r, column=age_col_idx).fill = PatternFill(
                        start_color=color, end_color=color, fill_type="solid"
                    )

        for c in range(1, ws.max_column + 1):
            max_len = max(
                len(str(ws.cell(row=r, column=c).value))
                if ws.cell(row=r, column=c).value else 0
                for r in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[get_column_letter(c)].width = min(max(8, max_len + 2), 50)

    wb.save(path)

def split_to_workbook(input_path, output_path, sheet_name, app_col):
    df = pd.read_excel(input_path, sheet_name=sheet_name, dtype=str)
    if app_col not in df.columns:
        raise KeyError(f"Missing column: {app_col}")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for app in sorted(df[app_col].fillna("Undefined").unique()):
            sub = df[df[app_col].fillna("Undefined") == app]
            sub.to_excel(writer, sheet_name=sanitize_sheet_name(app), index=False)

    style_workbook(pathlib.Path(output_path))
    print(f"Wrote workbook: {output_path}")

def main():
    p = argparse.ArgumentParser()
    p.add_argument("input")
    p.add_argument("-o", "--output", default="Vul_By_App.xlsx")
    p.add_argument("--sheet", default="Vulnerabilities")
    p.add_argument("--app-col", default="Application Name")
    args = p.parse_args()

    input_path = pathlib.Path(args.input)
    if not input_path.exists():
        print("Input file not found", file=sys.stderr)
        sys.exit(2)

    split_to_workbook(
        input_path,
        pathlib.Path(args.output),
        args.sheet,
        args.app_col
    )

if __name__ == "__main__":
    main()
